import json
import os

from django.views.generic import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http.response import JsonResponse, HttpResponse

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles.colors import BLACK
from openpyxl.styles import Border, Side

from inventory_control.models import InventoryCount, Inventory


class InventoryCountReport(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        params = request.GET
        filter_data = {}
        end_date = params.get('end_date', '')
        inventory_id = params.get('inventory_id', '')

        inventory_instance = Inventory.objects.get(id=inventory_id)
        inventory_counts = inventory_instance.inventory_records.all()

        for key, value in params.items():
            if value != '':
                if key == 'sku':
                    filter_data['product__sku'] = value
                if key == 'product_status':
                    filter_data['product_status'] = value
                if key == 'storage_type':
                    filter_data['storage_type'] = value


        inventory_counts = inventory_counts.filter(**filter_data)

        # os.chdir('/home/admin/server')

        wb = load_workbook('core/static/assets/report_templates/report.xlsx')
        ws = wb.active
        initial_row = 13

        border =  Border(
            left=Side(style='thin',color=BLACK),
            right=Side(style='thin',color=BLACK),
            top=Side(style='thin', color=BLACK),
            bottom=Side(style='thin',  color=BLACK)
        )
        
        ws['B4'] = inventory_instance.inventory_date

        for item in inventory_counts:
            ws.cell(row=initial_row, column=1).value = item.inventory.store.username
            ws.cell(row=initial_row, column=2).value = item.inventory.employee.first_name
            ws.cell(row=initial_row, column=3).value = item.created.date()
            ws.cell(row=initial_row, column=4).value = item.inventory.warehouse.name
            ws.cell(row=initial_row, column=5).value = item.storage_type.name
            ws.cell(row=initial_row, column=6).value = item.level
            ws.cell(row=initial_row, column=7).value = item.position
            ws.cell(row=initial_row, column=8).value = item.product.description
            ws.cell(row=initial_row, column=9).value = item.product.sku
            ws.cell(row=initial_row, column=10).value = item.amount
            ws.cell(row=initial_row, column=11).value = item.product_status.name
            ws.cell(row=initial_row, column=12).value = item.image.url if item.image else ''

            for column in range(1, 13):
                ws.cell(row=initial_row, column=column).border = border

            initial_row = initial_row + 1


        filename = 'Reporte de conteos.xlsx'
        response = HttpResponse(content_type='application/ms-excel')
        content = f'attachment; filename={filename}'
        response['Content-Disposition'] = content
        wb.save(response)
        return response